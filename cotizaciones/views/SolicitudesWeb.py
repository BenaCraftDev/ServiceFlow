import json
import csv
from decimal import Decimal
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.template.loader import render_to_string
from django.conf import settings
from ..models import *
from ..forms import *
from ..forms_empleados import *
from ..forms_prestamos import *
from home.decorators import requiere_admin, requiere_gerente_o_superior
from notificaciones.models import Notificacion
from notificaciones.utils import crear_notificacion
from home.models import PerfilEmpleado
from ..utils_mantenimiento import verificar_mantenimientos_materiales

@login_required
@requiere_gerente_o_superior
def obtener_servicios_categoria(request, categoria_id):
    """Obtener servicios de una categoría vía AJAX"""
    servicios = ServicioBase.objects.filter(
        categoria_id=categoria_id, 
        activo=True
    ).values('id', 'nombre', 'precio_base', 'unidad', 'es_parametrizable')
    
    return JsonResponse(list(servicios), safe=False)

@login_required
@requiere_gerente_o_superior
def obtener_parametros_servicio(request, servicio_id):
    """Obtener parámetros de un servicio vía AJAX"""
    try:
        parametros = ParametroServicio.objects.filter(
            servicio_id=servicio_id
        ).order_by('orden')
        
        parametros_data = []
        for param in parametros:
            param_dict = {
                'id': param.id,
                'nombre': param.nombre,
                'tipo': param.tipo,
                'requerido': param.requerido,
                'valor_por_defecto': param.valor_por_defecto or '',
                'orden': param.orden
            }
            
            if param.tipo == 'select' and param.opciones:
                param_dict['opciones_list'] = [opt.strip() for opt in param.opciones.split(',')]
            
            parametros_data.append(param_dict)
        
        return JsonResponse({
            'success': True,
            'parametros': parametros_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@requiere_gerente_o_superior
def lista_solicitudes_web(request):
    """
    Vista para listar todas las solicitudes web.
    Muestra solicitudes pendientes, en revisión y convertidas.
    """
    # Filtros
    estado_filtro = request.GET.get('estado', 'pendiente')
    busqueda = request.GET.get('busqueda', '').strip()
    
    # Query base
    solicitudes = Solicitud_Web.objects.all()
    
    # Aplicar filtro de estado
    if estado_filtro and estado_filtro != 'todas':
        solicitudes = solicitudes.filter(estado=estado_filtro)
    
    # Aplicar búsqueda
    if busqueda:
        from django.db.models import Q
        solicitudes = solicitudes.filter(
            Q(nombre_solicitante__icontains=busqueda) |
            Q(email_solicitante__icontains=busqueda) |
            Q(telefono_solicitante__icontains=busqueda) |
            Q(tipo_servicio_solicitado__icontains=busqueda) |
            Q(ubicacion_trabajo__icontains=busqueda)
        )
    
    # Ordenar por fecha más reciente
    solicitudes = solicitudes.order_by('-fecha_solicitud')
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page', 1)
    solicitudes_paginadas = paginator.get_page(page_number)
    
    # Estadísticas rápidas
    stats = {
        'pendientes': Solicitud_Web.objects.filter(estado='pendiente').count(),
        'en_revision': Solicitud_Web.objects.filter(estado='en_revision').count(),
        'convertidas': Solicitud_Web.objects.filter(estado='convertida').count(),
        'descartadas': Solicitud_Web.objects.filter(estado='descartada').count(),
        'urgentes': Solicitud_Web.objects.filter(
            estado='pendiente',
            fecha_solicitud__lt=timezone.now() - timezone.timedelta(days=2)
        ).count()
    }
    
    context = {
        'solicitudes': solicitudes_paginadas,
        'estado_filtro': estado_filtro,
        'busqueda': busqueda,
        'stats': stats,
    }
    
    return render(request, 'cotizaciones/solicitudes_web/lista.html', context)

@login_required
@requiere_gerente_o_superior
def detalle_solicitud_web(request, pk):
    """
    Vista detallada de una solicitud web.
    Permite revisar todos los datos y tomar acciones.
    """
    solicitud = get_object_or_404(Solicitud_Web, pk=pk)
    
    # Si la solicitud está pendiente, marcarla como en revisión
    if solicitud.estado == 'pendiente':
        solicitud.marcar_en_revision(request.user)
    
    # Buscar clientes similares por nombre o teléfono
    from django.db.models import Q
    clientes_similares = Cliente.objects.filter(
        Q(nombre__icontains=solicitud.nombre_solicitante) |
        Q(telefono__icontains=solicitud.telefono_solicitante) |
        Q(email__icontains=solicitud.email_solicitante if solicitud.email_solicitante else 'xxx')
    )[:10]
    
    # Tipos de trabajo disponibles
    tipos_trabajo = TipoTrabajo.objects.filter(activo=True)
    
    context = {
        'solicitud': solicitud,
        'clientes_similares': clientes_similares,
        'tipos_trabajo': tipos_trabajo,
    }
    
    return render(request, 'cotizaciones/solicitudes_web/detalle.html', context)

@login_required
@requiere_gerente_o_superior
def convertir_solicitud_web_a_cotizacion(request, pk):
    """
    Convierte una solicitud web en cotización formal.
    
    ✅ Permite seleccionar o crear cliente
    ✅ Asigna tipo de trabajo
    ✅ NO modifica datos existentes
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        solicitud = get_object_or_404(Solicitud_Web, pk=pk)
        
        # Verificar que no esté ya convertida
        if solicitud.estado == 'convertida':
            return JsonResponse({
                'success': False,
                'error': 'Esta solicitud ya fue convertida'
            }, status=400)
        
        # Obtener datos del formulario
        accion_cliente = request.POST.get('accion_cliente')  # 'nuevo' o 'existente'
        tipo_trabajo_id = request.POST.get('tipo_trabajo_id')
        
        # Validar tipo de trabajo
        tipo_trabajo = get_object_or_404(TipoTrabajo, pk=tipo_trabajo_id, activo=True)
        
        # ═══════════════════════════════════════════════════════
        # MANEJAR CLIENTE
        # ═══════════════════════════════════════════════════════
        with transaction.atomic():
            if accion_cliente == 'nuevo':
                # CREAR NUEVO CLIENTE con datos de la solicitud
                cliente = Cliente.objects.create(
                    nombre=solicitud.nombre_solicitante,
                    email=solicitud.email_solicitante,
                    telefono=solicitud.telefono_solicitante,
                    direccion=solicitud.ubicacion_trabajo
                )
                messages.success(request, f'Cliente creado: {cliente.nombre}')
                
            elif accion_cliente == 'existente':
                # USAR CLIENTE EXISTENTE (sin modificar sus datos)
                cliente_id = request.POST.get('cliente_id')
                cliente = get_object_or_404(Cliente, pk=cliente_id)
                messages.info(request, f'Usando cliente existente: {cliente.nombre}')
                
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar crear nuevo cliente o usar uno existente'
                }, status=400)
            
            # ═══════════════════════════════════════════════════════
            # CONVERTIR A COTIZACIÓN
            # ═══════════════════════════════════════════════════════
            cotizacion = solicitud.convertir_a_cotizacion(
                usuario=request.user,
                cliente=cliente,
                tipo_trabajo=tipo_trabajo
            )
            
            # Notificar éxito
            messages.success(
                request,
                f'¡Solicitud convertida exitosamente! Cotización #{cotizacion.numero}'
            )
            
            # Crear notificación
            crear_notificacion(
                usuario=request.user,
                tipo='success',
                titulo='Solicitud Web Convertida',
                mensaje=f'Solicitud de {solicitud.nombre_solicitante} convertida a cotización {cotizacion.numero}',
                url=f'/cotizaciones/{cotizacion.id}/editar/'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Cotización {cotizacion.numero} creada exitosamente',
                'cotizacion_id': cotizacion.id,
                'cotizacion_numero': cotizacion.numero,
                'redirect_url': f'/cotizaciones/{cotizacion.id}/editar/'
            })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al convertir: {str(e)}'
        }, status=500)

@login_required
@requiere_gerente_o_superior
def descartar_solicitud_web(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)
    
    try:
        solicitud = get_object_or_404(Solicitud_Web, pk=pk)
        
        if solicitud.estado == 'convertida':
            return JsonResponse({'success': False, 'error': 'No se puede descartar una solicitud ya convertida'}, status=400)
        
        motivo = request.POST.get('motivo', '').strip()
        solicitud.marcar_descartada(request.user, motivo)

        # --- LÓGICA DE ENVÍO DE EMAIL ---
        if solicitud.email_solicitante:
            subject = f"Actualización sobre su solicitud #{solicitud.id}"
            html_content = f"""
                <div style="font-family: sans-serif; color: #333;">
                    <h2>Hola, {solicitud.nombre_solicitante}.</h2>
                    <p>Le agradecemos su interés en nuestros servicios.</p>
                    <p>Lamentamos informarle que su solicitud de <strong>{solicitud.tipo_servicio_solicitado}</strong> ha sido desestimada en este momento.</p>
                    {f"<p><strong>Motivo:</strong> {motivo}</p>" if motivo else ""}
                    <p>Si tiene dudas, puede contactarnos directamente respondiendo a este correo.</p>
                    <br>
                    <p>Atentamente,<br>El equipo de Gestión.</p>
                </div>
            """
            # Llamamos a tu función existente
            enviar_email_con_reintentos(
                subject=subject,
                html_content=html_content,
                recipient_list=[solicitud.email_solicitante],
                fail_silently=True  # Para que el descarte funcione aunque el mail falle
            )

        messages.warning(request, f'Solicitud #{solicitud.id} descartada y cliente notificado.')
        return JsonResponse({'success': True, 'message': 'Solicitud descartada'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@requiere_gerente_o_superior
def agregar_nota_solicitud_web(request, pk):
    """
    Agrega una nota interna a una solicitud web.
    """
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Método no permitido'
        }, status=405)
    
    try:
        solicitud = get_object_or_404(Solicitud_Web, pk=pk)
        nota = request.POST.get('nota', '').strip()
        
        if not nota:
            return JsonResponse({
                'success': False,
                'error': 'La nota no puede estar vacía'
            }, status=400)
        
        # Agregar nota con timestamp y usuario
        timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
        usuario = request.user.get_full_name() or request.user.username
        nueva_nota = f"[{timestamp}] {usuario}:\n{nota}\n"
        
        if solicitud.notas_internas:
            solicitud.notas_internas = nueva_nota + "\n" + solicitud.notas_internas
        else:
            solicitud.notas_internas = nueva_nota
        
        solicitud.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Nota agregada'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@requiere_gerente_o_superior
def exportar_solicitudes_web(request):
    """
    Exporta solicitudes web a Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes Web"
    
    # Estilos
    header_fill = PatternFill(start_color="2575C0", end_color="2575C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Headers
    headers = [
        'ID', 'Fecha', 'Estado', 'Nombre', 'Email', 'Teléfono',
        'Servicio', 'Ubicación', 'Info Extra', 'Días Pendiente',
        'Procesada Por', 'Cotización'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    solicitudes = Solicitud_Web.objects.all().order_by('-fecha_solicitud')
    
    for row, sol in enumerate(solicitudes, 2):
        ws.cell(row=row, column=1, value=sol.id)
        ws.cell(row=row, column=2, value=sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=3, value=sol.get_estado_display())
        ws.cell(row=row, column=4, value=sol.nombre_solicitante)
        ws.cell(row=row, column=5, value=sol.email_solicitante or '')
        ws.cell(row=row, column=6, value=sol.telefono_solicitante)
        ws.cell(row=row, column=7, value=sol.tipo_servicio_solicitado)
        ws.cell(row=row, column=8, value=sol.ubicacion_trabajo)
        ws.cell(row=row, column=9, value=sol.informacion_adicional or '')
        ws.cell(row=row, column=10, value=sol.get_dias_pendiente() if sol.estado == 'pendiente' else '-')
        ws.cell(row=row, column=11, value=sol.procesada_por.username if sol.procesada_por else '')
        ws.cell(row=row, column=12, value=sol.cotizacion_generada.numero if sol.cotizacion_generada else '')
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 20
    ws.column_dimensions['L'].width = 15
    
    # Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=solicitudes_web_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response

@login_required
@requiere_gerente_o_superior
def ver_solicitudes_pendientes(request):
    """
    Vista para listar todas las solicitudes web pendientes
    """
    # Obtener solicitudes pendientes
    solicitudes = Cotizacion.objects.filter(
        estado='pedido'
    ).select_related(
        'cliente', 
        'tipo_trabajo'
    ).order_by('-fecha_creacion')
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page', 1)
    solicitudes_paginadas = paginator.get_page(page_number)
    
    context = {
        'solicitudes': solicitudes_paginadas,
        'total_solicitudes': solicitudes.count(),
    }
    
    return render(request, 'cotizaciones/solicitudes/lista_solicitudes.html', context)
