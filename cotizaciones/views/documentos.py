import io
import json
import csv
from decimal import Decimal
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg
from django.utils import timezone
from django.template.loader import render_to_string
from django.conf import settings
from ..models import *
from ..forms import *
from ..forms_empleados import *
from ..forms_prestamos import *
from home.decorators import requiere_admin, requiere_gerente_o_superior
from notificaciones.models import Notificacion
from notificaciones.utils import crear_notificacion
from home.models import PerfilEmpleado
from ..utils_mantenimiento import verificar_mantenimientos_materiales
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT



@login_required
@requiere_gerente_o_superior
def exportar_tipos_trabajo(request):
    """Exportar tipos de trabajo a Excel o CSV"""
    
    formato = request.GET.get('formato', 'excel')
    estado_filtro = request.GET.get('estado', '').strip()
    
    # Query
    tipos = TipoTrabajo.objects.all()
    
    if estado_filtro == 'activo':
        tipos = tipos.filter(activo=True)
    elif estado_filtro == 'inactivo':
        tipos = tipos.filter(activo=False)
    
    tipos = tipos.order_by('nombre')
    
    if formato == 'csv':
        # Exportar CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="tipos_trabajo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        # BOM para UTF-8
        response.write('\ufeff')
        
        writer = csv.writer(response)
        writer.writerow(['Nombre', 'Descripción', 'Estado', 'Cotizaciones Asociadas'])
        
        for tipo in tipos:
            writer.writerow([
                tipo.nombre,
                tipo.descripcion or '',
                'Activo' if tipo.activo else 'Inactivo',
                tipo.cotizacion_set.count()
            ])
        
        return response
    
    else:
        # Exportar Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Tipos de Trabajo"
        
        # Estilos
        header_fill = PatternFill(start_color='2575C0', end_color='2575C0', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Nombre', 'Descripción', 'Estado', 'Cotizaciones Asociadas']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_num, tipo in enumerate(tipos, 2):
            ws.cell(row=row_num, column=1, value=tipo.nombre).border = border
            ws.cell(row=row_num, column=2, value=tipo.descripcion or '').border = border
            ws.cell(row=row_num, column=3, value='Activo' if tipo.activo else 'Inactivo').border = border
            ws.cell(row=row_num, column=4, value=tipo.cotizacion_set.count()).border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Guardar
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tipos_trabajo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response

# === Cotizaciones ===

@login_required
@requiere_gerente_o_superior
def generar_pdf_cotizacion(request, pk):
    cotizacion = get_object_or_404(Cotizacion, pk=pk)
    
    # Crear el objeto HttpResponse con el content-type de PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Cotizacion_{cotizacion.numero}.pdf"'
    
    # Crear el buffer en memoria
    buffer = io.BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    # Contenedor para los elementos del PDF
    elements = []
    
    # Estilos personalizados
    styles = getSampleStyleSheet()
    
    # Estilo para encabezado empresa (centrado)
    empresa_style = ParagraphStyle(
        'EmpresaStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceBefore=0,
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )
    
    info_empresa_style = ParagraphStyle(
        'InfoEmpresaStyle', 
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceBefore=1,
        spaceAfter=1
    )
    
    # Estilo para título cotización (centrado)
    titulo_cot_style = ParagraphStyle(
        'TituloCotStyle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.black,
        alignment=TA_CENTER,
        spaceBefore=15,
        spaceAfter=15,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para información del cliente (alineado a la izquierda)
    cliente_style = ParagraphStyle(
        'ClienteStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.black,
        alignment=TA_LEFT,
        spaceBefore=3,
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )
    
    # Estilo para subtítulos de secciones
    seccion_style = ParagraphStyle(
        'SeccionStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.black,
        spaceBefore=15,
        spaceAfter=10,
        fontName='Helvetica-Bold',
        alignment=TA_LEFT
    )
    
    # Estilo para totales (alineado a la derecha)
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.black,
        alignment=TA_RIGHT,
        spaceBefore=3,
        spaceAfter=3,
        fontName='Helvetica-Bold'
    )
    
    # ENCABEZADO EMPRESA (CENTRADO)
    elements.append(Paragraph("JOSE E. ALVARADO N.", empresa_style))
    elements.append(Paragraph("SERVICIOS ELECTROMECANICOS", info_empresa_style))
    elements.append(Paragraph("INSTALACIÓN, MANTENCIÓN Y REPARACIÓN DE BOMBAS DE AGUA.", info_empresa_style))
    elements.append(Paragraph("SUPERFICIE Y SUMERGIBLES", info_empresa_style))
    elements.append(Paragraph("Pje. Santa Elisa 2437 Osorno", info_empresa_style))
    elements.append(Paragraph("TELEFONOS: 9-76193683/ EMAIL: seelmec@gmail.com", info_empresa_style))
    
    # LÍNEA DE SEPARACIÓN
    elements.append(Spacer(1, 10))
    line_table = Table([['_' * 80]], colWidths=[17*cm])
    line_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (0, 0), 12),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.black),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 10))
    
    # TÍTULO COTIZACIÓN (CENTRADO)
    elements.append(Paragraph(f"COTIZACIÓN   N°  {cotizacion.numero}", titulo_cot_style))
    
    # INFORMACIÓN DEL CLIENTE (ALINEADO A LA IZQUIERDA)
    elements.append(Paragraph(f"SEÑOR(ES): {cotizacion.get_nombre_cliente().upper()}", cliente_style))
    
    if cotizacion.representante or cotizacion.representante_nombre_respaldo:
        rep_nombre = cotizacion.get_nombre_representante()
        if rep_nombre:
            elements.append(Paragraph(f"ATENCIÓN: {rep_nombre.upper()}", cliente_style))

    elements.append(Paragraph(f"REFERENCIA: {cotizacion.referencia.upper()}", cliente_style))
    elements.append(Paragraph(f"LUGAR: {cotizacion.lugar.upper()}", cliente_style))
    
    # SUBTÍTULO DESCRIPCIÓN
    elements.append(Paragraph("A.- DESCRIPCIÓN DE TRABAJOS, DETALLE Y VALORIZACIÓN.", seccion_style))
    elements.append(Spacer(1, 10))
    
    # OBTENER ITEMS
    try:
        from .models import ItemServicio, ItemMaterial, ItemManoObra
        
        items_servicio = list(ItemServicio.objects.filter(cotizacion=cotizacion).select_related('servicio'))
        items_material = list(ItemMaterial.objects.filter(cotizacion=cotizacion).select_related('material'))
        items_mano_obra = list(ItemManoObra.objects.filter(cotizacion=cotizacion))
        
    except Exception as e:
        items_servicio = []
        items_material = []
        items_mano_obra = []
    
    # TABLA DE SERVICIOS
    if items_servicio:
        servicios_data = [['DESCRIPCIÓN DEL TRABAJO', 'CANTIDAD', 'PRECIO UNIT.', 'SUBTOTAL']]
        
        for item in items_servicio:
            if item.descripcion_personalizada:
                descripcion = item.descripcion_personalizada
            else:
                descripcion = str(item.servicio)
            
            cantidad = f"{item.cantidad} {item.servicio.unidad if item.servicio else 'UND'}"
            precio = f"${int(item.precio_unitario):,}".replace(',', '.')
            subtotal = f"${int(item.subtotal):,}".replace(',', '.')
            
            servicios_data.append([descripcion, cantidad, precio, subtotal])
        
        servicios_table = Table(servicios_data, colWidths=[9*cm, 2.5*cm, 2.5*cm, 3*cm])
        servicios_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(servicios_table)
        elements.append(Spacer(1, 10))
    
    # TABLA DE MATERIALES
    if items_material:
        materiales_data = [['MATERIAL', 'CANTIDAD', 'PRECIO UNIT.', 'SUBTOTAL']]
        
        for item in items_material:
            if item.descripcion_personalizada:
                descripcion = item.descripcion_personalizada
            elif item.material:
                descripcion = f"{item.material.codigo} - {item.material.nombre}"
            else:
                descripcion = "Material sin especificar"
            
            unidad = item.material.unidad if item.material else "UND"
            cantidad = f"{item.cantidad} {unidad}"
            precio = f"${int(item.precio_unitario):,}".replace(',', '.')
            subtotal = f"${int(item.subtotal):,}".replace(',', '.')
            
            materiales_data.append([descripcion, cantidad, precio, subtotal])
        
        materiales_table = Table(materiales_data, colWidths=[9*cm, 2.5*cm, 2.5*cm, 3*cm])
        materiales_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(materiales_table)
        elements.append(Spacer(1, 10))
    
    # TABLA DE MANO DE OBRA
    if items_mano_obra:
        mano_obra_data = [['MANO DE OBRA', 'HORAS', 'PRECIO/HORA', 'SUBTOTAL']]
        
        for item in items_mano_obra:
            descripcion = item.descripcion
            horas = f"{item.horas}"
            precio_hora = f"${int(item.precio_hora):,}".replace(',', '.')
            subtotal = f"${int(item.subtotal):,}".replace(',', '.')
            
            mano_obra_data.append([descripcion, horas, precio_hora, subtotal])
        
        mano_obra_table = Table(mano_obra_data, colWidths=[9*cm, 2.5*cm, 2.5*cm, 3*cm])
        mano_obra_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(mano_obra_table)
        elements.append(Spacer(1, 15))
    
    # SECCIÓN DE TOTALES
    elements.append(Spacer(1, 20))
    
    # Crear tabla con conceptos a la izquierda y valores a la derecha
    totales_data = [
        ['VALOR TOTAL TRABAJOS', f"${int(cotizacion.subtotal_servicios):,}".replace(',', '.')],
        ['MATERIALES', f"${int(cotizacion.subtotal_materiales):,}".replace(',', '.')],
        ['MANO DE OBRA', f"${int(cotizacion.subtotal_mano_obra):,}".replace(',', '.')],
        ['GASTOS DE TRASLADO', f"${int(cotizacion.gastos_traslado):,}".replace(',', '.')],
        ['', ''],  # Espacio
        ['VALOR NETO', f"${int(cotizacion.valor_neto):,}".replace(',', '.')],
        ['VALOR IVA (19%)', f"${int(cotizacion.valor_iva):,}".replace(',', '.')],
        ['VALOR TOTAL', f"${int(cotizacion.valor_total):,}".replace(',', '.')]
    ]
    
    totales_table = Table(totales_data, colWidths=[10*cm, 7*cm])
    totales_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Conceptos a la izquierda
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Valores a la derecha
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -3), 'Helvetica'),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),  # Últimas dos filas en negrita
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('LINEABOVE', (0, -2), (-1, -2), 1, colors.black),  # Línea antes del valor neto
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),  # Línea más gruesa antes del total
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    elements.append(totales_table)
    
    
    # OBSERVACIONES
    elements.append(Spacer(1, 25))
    if cotizacion.observaciones:
        elements.append(Paragraph(f"NOTA: {cotizacion.observaciones}", styles['Normal']))
    else:
        elements.append(Paragraph("NOTA: Sin observaciones adicionales.", styles['Normal']))
    
    # SECCIÓN DE FIRMAS
    elements.append(Spacer(1, 20))
    
    # Crear tabla para firmas con espaciado adecuado
    firmas_data = [
        ['SALUDA ATTE.       JOSE E. ALVARADO N.', '', ''],
        ['', '', ''],
        ['FIRMA:_________________', '', 'NOMBRE Y RUT:_________________'],
        ['', '', 'ACEPTADO CLIENTE']
    ]
    
    firmas_table = Table(firmas_data, colWidths=[7*cm, 3*cm, 7*cm])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),     # SALUDA ATTE a la izquierda
        ('ALIGN', (0, 2), (0, 2), 'LEFT'),     # FIRMA a la izquierda
        ('ALIGN', (2, 2), (2, 2), 'CENTER'),   # NOMBRE Y RUT centrado
        ('ALIGN', (2, 3), (2, 3), 'CENTER'),   # ACEPTADO CLIENTE centrado
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    elements.append(firmas_table)
    
    # FECHA (alineada a la izquierda)
    elements.append(Spacer(1, 30))
    
    # Formatear fecha como en la imagen: "OSORNO 20 DE SEPTIEMBRE DE 2022"
    meses = {
        1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO', 6: 'JUNIO',
        7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
    }
    
    dia = cotizacion.fecha_creacion.day
    mes = meses[cotizacion.fecha_creacion.month]
    año = cotizacion.fecha_creacion.year
    
    fecha_formateada = f"OSORNO {dia} DE {mes} DE {año}"
    
    elements.append(Paragraph(fecha_formateada, 
                            ParagraphStyle('FechaStyle', parent=styles['Normal'], 
                                         fontSize=10, alignment=TA_LEFT, 
                                         fontName='Helvetica')))
    
    # Construir el PDF
    doc.build(elements)
    
    # Obtener el valor del buffer y escribirlo a la respuesta
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

@login_required
@requiere_gerente_o_superior
def exportar_cotizaciones(request):
    """Exportar cotizaciones a Excel o CSV"""
    formato = request.GET.get('formato', 'excel')
    
    cotizaciones = Cotizacion.objects.select_related('cliente', 'tipo_trabajo').order_by('-fecha_creacion')
    
    # Aplicar filtros
    busqueda = request.GET.get('busqueda', '')
    estado = request.GET.get('estado', '')
    cliente_id = request.GET.get('cliente', '')
    
    if busqueda:
        cotizaciones = cotizaciones.filter(
            Q(numero__icontains=busqueda) |
            Q(cliente__nombre__icontains=busqueda) |
            Q(referencia__icontains=busqueda)
        )
    
    if estado:
        cotizaciones = cotizaciones.filter(estado=estado)
        
    if cliente_id:
        cotizaciones = cotizaciones.filter(cliente_id=cliente_id)
    
    if formato == 'csv':
        return exportar_cotizaciones_csv(cotizaciones)
    else:
        return exportar_cotizaciones_excel(cotizaciones)

def exportar_cotizaciones_csv(cotizaciones):
    """Exportar cotizaciones a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="cotizaciones.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'Número', 'Cliente', 'Referencia', 'Lugar', 'Tipo Trabajo',
        'Fecha Creación', 'Estado', 'Valor Neto', 'IVA', 'Valor Total'
    ])
    
    for cot in cotizaciones:
        writer.writerow([
            cot.numero,
            cot.cliente.nombre if cot.cliente else 'Sin cliente',
            cot.referencia,
            cot.lugar,
            cot.tipo_trabajo.nombre if cot.tipo_trabajo else 'Sin tipo',
            cot.fecha_creacion.strftime('%d/%m/%Y'),
            cot.get_estado_display(),
            float(cot.valor_neto),
            float(cot.valor_iva),
            float(cot.valor_total)
        ])
    
    return response

def exportar_cotizaciones_excel(cotizaciones):
    """Exportar cotizaciones a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Número', 'Cliente', 'Referencia', 'Lugar', 'Tipo Trabajo', 
               'Fecha Creación', 'Estado', 'Valor Neto', 'IVA', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, cot in enumerate(cotizaciones, 2):
        ws.cell(row=row, column=1, value=cot.numero).border = border
        ws.cell(row=row, column=2, value=cot.cliente.nombre if cot.cliente else 'Sin cliente').border = border
        ws.cell(row=row, column=3, value=cot.referencia).border = border
        ws.cell(row=row, column=4, value=cot.lugar).border = border
        ws.cell(row=row, column=5, value=cot.tipo_trabajo.nombre if cot.tipo_trabajo else 'Sin tipo').border = border
        ws.cell(row=row, column=6, value=cot.fecha_creacion.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=7, value=cot.get_estado_display()).border = border
        ws.cell(row=row, column=8, value=float(cot.valor_neto)).border = border
        ws.cell(row=row, column=9, value=float(cot.valor_iva)).border = border
        ws.cell(row=row, column=10, value=float(cot.valor_total)).border = border
    
    # Ajustar ancho de columnas
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cotizaciones.xlsx"'
    wb.save(response)
    
    return response
    """Exportar cotizaciones a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotizaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Número', 'Cliente', 'Referencia', 'Lugar', 'Tipo Trabajo', 
               'Fecha Creación', 'Estado', 'Valor Neto', 'IVA', 'Valor Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, cot in enumerate(cotizaciones, 2):
        ws.cell(row=row, column=1, value=cot.numero).border = border
        ws.cell(row=row, column=2, value=cot.cliente.nombre).border = border
        ws.cell(row=row, column=3, value=cot.referencia).border = border
        ws.cell(row=row, column=4, value=cot.lugar).border = border
        ws.cell(row=row, column=5, value=cot.tipo_trabajo.nombre).border = border
        ws.cell(row=row, column=6, value=cot.fecha_creacion.strftime('%d/%m/%Y')).border = border
        ws.cell(row=row, column=7, value=cot.get_estado_display()).border = border
        ws.cell(row=row, column=8, value=float(cot.valor_neto)).border = border
        ws.cell(row=row, column=9, value=float(cot.valor_iva)).border = border
        ws.cell(row=row, column=10, value=float(cot.valor_total)).border = border
    
    # Ajustar ancho de columnas
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cotizaciones.xlsx"'
    wb.save(response)
    
    return response

# === Clientes ===

@login_required
@requiere_gerente_o_superior
def exportar_clientes(request):
    """Exportar clientes a Excel o CSV"""
    formato = request.GET.get('formato', 'excel')
    
    # Aplicar los mismos filtros que en la vista principal
    clientes = Cliente.objects.all().order_by('nombre')
    busqueda = request.GET.get('busqueda', '')
    if busqueda:
        clientes = clientes.filter(
            Q(nombre__icontains=busqueda) |
            Q(rut__icontains=busqueda) |
            Q(email__icontains=busqueda)
        )
    
    if formato == 'csv':
        return exportar_clientes_csv(clientes)
    else:
        return exportar_clientes_excel(clientes)

def exportar_clientes_csv(clientes):
    """Exportar clientes a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="clientes.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'Nombre', 'RUT', 'Dirección', 'Teléfono', 'Email', 'Fecha Creación'
    ])
    
    for cliente in clientes:
        # Obtener representantes
        representantes = ', '.join([r.nombre for r in cliente.representantes.all()])
        
        writer.writerow([
            cliente.nombre,
            cliente.rut or '',
            cliente.direccion or '',
            cliente.telefono or '',
            cliente.email or '',
            cliente.fecha_creacion.strftime('%d/%m/%Y'),
        ])
    
    return response

def exportar_clientes_excel(clientes):
    """Exportar clientes a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Nombre', 'Representantes', 'RUT', 'Dirección', 'Teléfono', 'Email', 'Fecha Creación']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, cliente in enumerate(clientes, 2):
        # Obtener representantes
        representantes = ', '.join([r.nombre for r in cliente.representantes.all()])
        
        ws.cell(row=row, column=1, value=cliente.nombre).border = border
        ws.cell(row=row, column=2, value=representantes or '-').border = border
        ws.cell(row=row, column=3, value=cliente.rut or '').border = border
        ws.cell(row=row, column=4, value=cliente.direccion or '').border = border
        ws.cell(row=row, column=5, value=cliente.telefono or '').border = border
        ws.cell(row=row, column=6, value=cliente.email or '').border = border
        ws.cell(row=row, column=7, value=cliente.fecha_creacion.strftime('%d/%m/%Y')).border = border
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    
    return response

# === Servicios ===

@login_required
@requiere_gerente_o_superior
def exportar_servicios(request):
    """Exportar servicios a Excel o CSV"""
    formato = request.GET.get('formato', 'excel')
    
    servicios = ServicioBase.objects.select_related('categoria').order_by('categoria__nombre', 'nombre')
    categoria_filtro = request.GET.get('categoria', '')
    if categoria_filtro:
        servicios = servicios.filter(categoria_id=categoria_filtro)
    
    if formato == 'csv':
        return exportar_servicios_csv(servicios)
    else:
        return exportar_servicios_excel(servicios)

def exportar_servicios_csv(servicios):
    """Exportar servicios a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="servicios.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'Categoría', 'Nombre', 'Descripción', 'Precio Base', 
        'Unidad', 'Parametrizable', 'Estado'
    ])
    
    for servicio in servicios:
        writer.writerow([
            servicio.categoria.nombre,
            servicio.nombre,
            servicio.descripcion,
            float(servicio.precio_base),
            servicio.unidad,
            'Sí' if servicio.es_parametrizable else 'No',
            'Activo' if servicio.activo else 'Inactivo'
        ])
    
    return response

def exportar_servicios_excel(servicios):
    """Exportar servicios a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Categoría', 'Nombre', 'Descripción', 'Precio Base', 'Unidad', 'Parametrizable', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, servicio in enumerate(servicios, 2):
        ws.cell(row=row, column=1, value=servicio.categoria.nombre).border = border
        ws.cell(row=row, column=2, value=servicio.nombre).border = border
        ws.cell(row=row, column=3, value=servicio.descripcion).border = border
        ws.cell(row=row, column=4, value=float(servicio.precio_base)).border = border
        ws.cell(row=row, column=5, value=servicio.unidad).border = border
        ws.cell(row=row, column=6, value='Sí' if servicio.es_parametrizable else 'No').border = border
        ws.cell(row=row, column=7, value='Activo' if servicio.activo else 'Inactivo').border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="servicios.xlsx"'
    wb.save(response)
    
    return response

# === Materiales ===

@login_required
@requiere_gerente_o_superior
def exportar_materiales(request):
    """Exportar materiales a Excel o CSV"""
    formato = request.GET.get('formato', 'excel')
    
    materiales = Material.objects.all().order_by('categoria', 'nombre')
    busqueda = request.GET.get('busqueda', '')
    categoria_filtro = request.GET.get('categoria', '')
    
    if busqueda:
        materiales = materiales.filter(
            Q(nombre__icontains=busqueda) |
            Q(codigo__icontains=busqueda) |
            Q(descripcion__icontains=busqueda)
        )
    
    if categoria_filtro:
        materiales = materiales.filter(categoria=categoria_filtro)
    
    if formato == 'csv':
        return exportar_materiales_csv(materiales)
    else:
        return exportar_materiales_excel(materiales)

def exportar_materiales_csv(materiales):
    """Exportar materiales a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="materiales.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'Código', 'Nombre', 'Descripción', 'Categoría', 
        'Precio Unitario', 'Unidad', 'Estado'
    ])
    
    for material in materiales:
        writer.writerow([
            material.codigo,
            material.nombre,
            material.descripcion or '',
            material.categoria or '',
            float(material.precio_unitario),
            material.unidad,
            'Activo' if material.activo else 'Inactivo'
        ])
    
    return response

def exportar_materiales_excel(materiales):
    """Exportar materiales a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materiales"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Código', 'Nombre', 'Descripción', 'Categoría', 'Precio Unitario', 'Unidad', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row, material in enumerate(materiales, 2):
        ws.cell(row=row, column=1, value=material.codigo).border = border
        ws.cell(row=row, column=2, value=material.nombre).border = border
        ws.cell(row=row, column=3, value=material.descripcion or '').border = border
        ws.cell(row=row, column=4, value=material.categoria.nombre if material.categoria else '').border = border
        ws.cell(row=row, column=5, value=float(material.precio_unitario)).border = border
        ws.cell(row=row, column=6, value=str(material.unidad) if material.unidad else '').border = border
        ws.cell(row=row, column=7, value='Activo' if material.activo else 'Inactivo').border = border
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="materiales.xlsx"'
    wb.save(response)
    
    return response

# === Trabajos ===

@login_required
@requiere_gerente_o_superior
def exportar_trabajos(request):
    """Exportar trabajos en progreso a Excel o CSV"""
    from django.db.models import Count, Sum, Q
    
    formato = request.GET.get('formato', 'excel')
    
    # Filtros
    filtro_estado = request.GET.get('estado', '')
    filtro_empleado = request.GET.get('empleado', '')
    busqueda = request.GET.get('busqueda', '')
    
    # Obtener cotizaciones aprobadas con trabajos
    cotizaciones_query = Cotizacion.objects.filter(
        estado='aprobada'
    ).annotate(
        total_trabajos=Count('trabajos_empleados')
    ).filter(total_trabajos__gt=0).select_related('cliente', 'tipo_trabajo')
    
    # Recopilar todos los trabajos
    trabajos_list = []
    
    for cot in cotizaciones_query:
        trabajos_q = cot.trabajos_empleados.select_related(
            'empleado__user',
            'item_mano_obra'
        ).all()
        
        # Aplicar filtros
        if filtro_estado:
            trabajos_q = trabajos_q.filter(estado=filtro_estado)
        
        if filtro_empleado:
            trabajos_q = trabajos_q.filter(empleado_id=filtro_empleado)
        
        for trabajo in trabajos_q:
            # Filtro de búsqueda
            if busqueda:
                busqueda_lower = busqueda.lower()
                if not (busqueda_lower in cot.numero.lower() or 
                       busqueda_lower in cot.get_nombre_cliente().lower() or 
                       busqueda_lower in trabajo.item_mano_obra.descripcion.lower()):
                    continue
            
            # Calcular progreso
            progreso = 0
            if trabajo.estado == 'completado':
                progreso = 100
            elif trabajo.estado == 'en_progreso' and trabajo.horas_trabajadas > 0:
                horas_item = trabajo.item_mano_obra.horas
                if horas_item > 0:
                    progreso = min((float(trabajo.horas_trabajadas) / float(horas_item)) * 100, 99)
                else:
                    progreso = 50
            
            trabajos_list.append({
                'cotizacion': cot,
                'trabajo': trabajo,
                'progreso': progreso
            })
    
    if formato == 'csv':
        return exportar_trabajos_csv(trabajos_list)
    else:
        return exportar_trabajos_excel(trabajos_list)

def exportar_trabajos_csv(trabajos_list):
    """Exportar trabajos a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="seguimiento_trabajos.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    writer.writerow([
        'Cotización', 'Cliente', 'Lugar', 'Trabajo', 'Empleado', 'Cargo',
        'Estado', 'Progreso (%)', 'Horas Trabajadas', 'Precio/Hora',
        'Fecha Inicio', 'Fecha Fin', 'Observaciones'
    ])
    
    for item in trabajos_list:
        cot = item['cotizacion']
        trabajo = item['trabajo']
        progreso = item['progreso']
        
        writer.writerow([
            cot.numero,
            cot.get_nombre_cliente(),
            cot.lugar,
            trabajo.item_mano_obra.descripcion,
            trabajo.empleado.nombre_completo,
            trabajo.empleado.get_cargo_display(),
            trabajo.get_estado_display(),
            f"{progreso:.0f}",
            f"{float(trabajo.horas_trabajadas):.1f}",
            f"{float(trabajo.item_mano_obra.precio_hora):,.0f}",
            trabajo.fecha_inicio.strftime('%d/%m/%Y %H:%M') if trabajo.fecha_inicio else 'No iniciado',
            trabajo.fecha_fin.strftime('%d/%m/%Y %H:%M') if trabajo.fecha_fin else 'En curso' if trabajo.estado == 'en_progreso' else '-',
            trabajo.observaciones_empleado or '-'
        ])
    
    return response

def exportar_trabajos_excel(trabajos_list):
    """Exportar trabajos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Seguimiento Trabajos"
    
    # Estilos
    header_fill = PatternFill(start_color="2575C0", end_color="2575C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Estado fills
    completado_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    en_progreso_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    pendiente_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    
    # Encabezados
    headers = [
        'Cotización', 'Cliente', 'Lugar', 'Trabajo', 'Empleado', 'Cargo',
        'Estado', 'Progreso (%)', 'Horas Trabajadas', 'Precio/Hora',
        'Fecha Inicio', 'Fecha Fin', 'Observaciones'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Datos
    for row_idx, item in enumerate(trabajos_list, 2):
        cot = item['cotizacion']
        trabajo = item['trabajo']
        progreso = item['progreso']
        
        # Determinar fill según estado
        if trabajo.estado == 'completado':
            row_fill = completado_fill
        elif trabajo.estado == 'en_progreso':
            row_fill = en_progreso_fill
        else:
            row_fill = pendiente_fill
        
        # Datos de la fila
        datos = [
            cot.numero,
            cot.get_nombre_cliente(),
            cot.lugar,
            trabajo.item_mano_obra.descripcion,
            trabajo.empleado.nombre_completo,
            trabajo.empleado.get_cargo_display(),
            trabajo.get_estado_display(),
            f"{progreso:.0f}%",
            f"{float(trabajo.horas_trabajadas):.1f}",
            f"${float(trabajo.item_mano_obra.precio_hora):,.0f}",
            trabajo.fecha_inicio.strftime('%d/%m/%Y %H:%M') if trabajo.fecha_inicio else 'No iniciado',
            trabajo.fecha_fin.strftime('%d/%m/%Y %H:%M') if trabajo.fecha_fin else ('En curso' if trabajo.estado == 'en_progreso' else '-'),
            trabajo.observaciones_empleado or '-'
        ]
        
        for col_idx, valor in enumerate(datos, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_idx == 7:  # Columna de estado
                cell.fill = row_fill
                cell.font = Font(bold=True)
    
    # Ajustar ancho de columnas
    column_widths = [15, 25, 20, 30, 25, 15, 15, 12, 15, 12, 18, 18, 40]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Establecer altura de fila del encabezado
    ws.row_dimensions[1].height = 30
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="seguimiento_trabajos.xlsx"'
    wb.save(response)
    
    return response


